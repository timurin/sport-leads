"""Thin shared tabular file parse / validate helpers (roadmap 4.5.1.1, ADR-020).

Section adapters own SoT mapping; this module only normalizes tables and envelopes.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable, Mapping, Sequence
from typing import Any, Literal

from openpyxl import load_workbook

from app.schemas.file_io import FileIoDryRunEnvelope, FileIoRowError, ParsedTable

RowValidator = Callable[[int, dict[str, str | None]], list[FileIoRowError]]

_CSV_SUFFIXES = (".csv", ".tsv", ".txt")
_XLSX_SUFFIXES = (".xlsx", ".xlsm")


class FileIoParseError(ValueError):
    """File cannot be parsed as a tabular CSV/XLSX payload."""


def detect_tabular_format(
    *,
    filename: str | None = None,
    content_type: str | None = None,
) -> Literal["csv", "xlsx"]:
    name = (filename or "").strip().lower()
    ctype = (content_type or "").strip().lower()

    if any(name.endswith(s) for s in _XLSX_SUFFIXES):
        return "xlsx"
    if any(name.endswith(s) for s in _CSV_SUFFIXES):
        return "csv"
    if "spreadsheetml" in ctype or "excel" in ctype:
        return "xlsx"
    if "csv" in ctype or "text/plain" in ctype or "tab-separated" in ctype:
        return "csv"
    raise FileIoParseError(
        "Cannot detect tabular format; provide .csv/.xlsx filename or content type"
    )


def normalize_header(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_tabular_bytes(
    data: bytes,
    *,
    filename: str | None = None,
    content_type: str | None = None,
    source_format: Literal["csv", "xlsx"] | None = None,
    sheet_name: str | None = None,
) -> ParsedTable:
    if not data:
        raise FileIoParseError("Empty file")

    fmt = source_format or detect_tabular_format(
        filename=filename, content_type=content_type
    )
    if fmt == "csv":
        return _parse_csv(data)
    return _parse_xlsx(data, sheet_name=sheet_name)


def require_columns(
    headers: Sequence[str],
    required: Sequence[str],
) -> list[FileIoRowError]:
    present = {h.strip().lower() for h in headers if h.strip()}
    errors: list[FileIoRowError] = []
    for col in required:
        key = col.strip().lower()
        if key not in present:
            errors.append(
                FileIoRowError(
                    row_number=0,
                    column=col,
                    code="missing_column",
                    message=f"Required column '{col}' is missing",
                )
            )
    return errors


def validate_rows(
    rows: Sequence[Mapping[str, str | None]],
    validators: Sequence[RowValidator],
) -> list[FileIoRowError]:
    errors: list[FileIoRowError] = []
    for index, row in enumerate(rows, start=1):
        as_dict = dict(row)
        for validator in validators:
            errors.extend(validator(index, as_dict))
    return errors


def build_dry_run_envelope(
    *,
    total_rows: int,
    errors: Sequence[FileIoRowError],
    preview: Sequence[Mapping[str, Any]] | None = None,
    preview_limit: int = 20,
    dry_run: bool = True,
) -> FileIoDryRunEnvelope:
    if total_rows < 0:
        raise ValueError("total_rows must be >= 0")
    if preview_limit < 0:
        raise ValueError("preview_limit must be >= 0")

    data_errors = [e for e in errors if e.row_number > 0]
    error_row_numbers = {e.row_number for e in data_errors}
    error_rows = len(error_row_numbers)
    # Header/file errors (row_number == 0) block commit even with zero data rows.
    header_blocked = any(e.row_number == 0 for e in errors)
    valid_rows = max(total_rows - error_rows, 0)
    preview_rows = [dict(item) for item in (preview or [])[:preview_limit]]
    can_commit = (not header_blocked) and error_rows == 0 and total_rows >= 0
    # Empty file with no header errors: can_commit True only if caller allows;
    # keep True when there are zero errors (including empty).
    return FileIoDryRunEnvelope(
        dry_run=dry_run,
        total_rows=total_rows,
        valid_rows=valid_rows,
        error_rows=error_rows,
        errors=list(errors),
        preview=preview_rows,
        can_commit=can_commit,
    )


def remap_row(
    row: Mapping[str, str | None],
    column_map: Mapping[str, str],
) -> dict[str, str | None]:
    """Map source header → canonical key. Unmapped columns are dropped."""
    by_lower = {str(k).strip().lower(): v for k, v in row.items()}
    out: dict[str, str | None] = {}
    for source, target in column_map.items():
        out[target] = by_lower.get(source.strip().lower())
    return out


def render_csv_bytes(
    headers: Sequence[str],
    rows: Sequence[Mapping[str, object | None]],
) -> bytes:
    """UTF-8 CSV with BOM for Excel-friendly open.

    Quote all non-empty fields so Excel (RU locale) does not re-split cells
    that contain `;` when it auto-detects semicolon as the column delimiter.
    """
    buffer = io.StringIO()
    writer = csv.writer(
        buffer,
        lineterminator="\n",
        quoting=csv.QUOTE_NONNUMERIC,
    )
    writer.writerow(list(headers))
    for row in rows:
        writer.writerow(
            [
                ""
                if row.get(header) is None
                else str(row.get(header))
                for header in headers
            ]
        )
    # Excel tip: honor comma delimiter even in semicolon-default locales.
    return ("\ufeffsep=,\n" + buffer.getvalue()).encode("utf-8")


def render_xlsx_bytes(
    headers: Sequence[str],
    rows: Sequence[Mapping[str, object | None]],
    *,
    sheet_name: str = "Sheet1",
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_name[:31] or "Sheet1"
    sheet.append(list(headers))
    for row in rows:
        sheet.append(
            [
                ""
                if row.get(header) is None
                else str(row.get(header))
                for header in headers
            ]
        )
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _detect_csv_delimiter(text: str) -> str:
    """Pick CSV delimiter without trusting Sniffer alone.

    Excel RU exports often use `;`. Sniffer fails or prefers `,` when cell
    values contain commas (e.g. product model names), collapsing the header
    into a single column.
    """
    first = ""
    for line in text.splitlines():
        if line.strip():
            first = line
            break
    if not first:
        return ","

    counts = {
        ";": first.count(";"),
        ",": first.count(","),
        "\t": first.count("\t"),
    }
    # Prefer the delimiter that splits the header into the most fields.
    best = max(counts.items(), key=lambda item: item[1])
    if best[1] > 0:
        return best[0]

    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        return ","


def _parse_csv(data: bytes) -> ParsedTable:
    text = _decode_text(data)
    # Skip Excel locale hints like "sep=," / "sep=;" (may appear after BOM).
    lines = text.splitlines()
    while lines and lines[0].strip().lower().startswith("sep="):
        lines.pop(0)
    text = "\n".join(lines)
    if not text.strip():
        raise FileIoParseError("Empty file")

    delimiter = _detect_csv_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matrix = [[normalize_header(cell) for cell in row] for row in reader]
    headers, rows = _matrix_to_table(matrix)
    return ParsedTable(headers=headers, rows=rows, source_format="csv")


def _parse_xlsx(data: bytes, *, sheet_name: str | None) -> ParsedTable:
    try:
        workbook = load_workbook(
            filename=io.BytesIO(data),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # noqa: BLE001 — surface as parse error
        raise FileIoParseError(f"Invalid XLSX file: {exc}") from exc

    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise FileIoParseError(
                    f"Sheet '{sheet_name}' not found; available: {workbook.sheetnames}"
                )
            sheet = workbook[sheet_name]
            resolved_name = sheet_name
        else:
            sheet = workbook[workbook.sheetnames[0]]
            resolved_name = workbook.sheetnames[0]

        matrix: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            matrix.append([normalize_header(cell) for cell in row])
    finally:
        workbook.close()

    headers, rows = _matrix_to_table(matrix)
    return ParsedTable(
        headers=headers,
        rows=rows,
        source_format="xlsx",
        sheet_name=resolved_name,
    )


def _matrix_to_table(
    matrix: list[list[str]],
) -> tuple[list[str], list[dict[str, str | None]]]:
    while matrix and all(not cell for cell in matrix[0]):
        matrix.pop(0)
    while matrix and all(not cell for cell in matrix[-1]):
        matrix.pop()

    if not matrix:
        raise FileIoParseError("File has no header row")

    raw_headers = matrix[0]
    if not any(raw_headers):
        raise FileIoParseError("File has no header row")

    headers = _unique_headers(raw_headers)
    width = len(headers)
    rows: list[dict[str, str | None]] = []
    for raw in matrix[1:]:
        padded = list(raw) + [""] * max(0, width - len(raw))
        padded = padded[:width]
        if all(not cell for cell in padded):
            continue
        rows.append(
            {
                headers[i]: (padded[i] if padded[i] != "" else None)
                for i in range(width)
            }
        )
    return headers, rows


def _unique_headers(raw: Sequence[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for index, name in enumerate(raw):
        base = name if name else f"column_{index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        out.append(base if count == 0 else f"{base}_{count + 1}")
    return out


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")
